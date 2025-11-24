# -*- coding: utf-8 -*-
"""
core/excel.py
Geração de Excel consolidado com múltiplas lojas lado-a-lado.
v4.9.1 - Adicionado suporte para N/A amarelo quando preço existe mas não é calculável
"""
from pathlib import Path
from typing import List, Dict, Optional
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .feed import FeedProduct


class ExcelBuilder:
    """
    Construtor de Excel multi-loja.
    
    Layout:
    | ID | Título | Ref Feed | Preço Feed | Loja1 Preço | Loja1 Dif% | Loja1 URL | Loja2 Preço | Loja2 Dif% | ...
    """
    
    def __init__(self, store_names: List[str]):
        """
        Args:
            store_names: Lista de nomes das lojas (ex: ["wrs", "omniaracing"])
        """
        self.store_names = store_names
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Comparador"
        
        # Estilos
        self._setup_styles()
    
    def _setup_styles(self):
        """Define estilos reutilizáveis"""
        # Header
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Borders
        thin_border = Side(border_style="thin", color="CCCCCC")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Formatação condicional
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # NOVO: Amarelo para N/A (preço existe mas não calculável)
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    def _create_headers(self):
        """Cria linha de headers"""
        # Headers fixos
        headers = ["ID", "Título", "Ref Feed", "Preço Feed"]
        
        # Headers por loja (3 colunas cada)
        for store_name in self.store_names:
            # Capitalizar nome da loja
            store_display = store_name.replace("_", " ").title()
            headers.extend([
                f"{store_display} Preço",
                f"{store_display} Dif%",
                f"{store_display} URL"
            ])
        
        # Adicionar headers
        self.ws.append(headers)
        
        # Formatar headers
        for col in range(1, len(headers) + 1):
            cell = self.ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_align
            cell.border = self.border
        
        # Ajustar larguras
        self._set_column_widths()
    
    def _set_column_widths(self):
        """Define larguras das colunas"""
        # Colunas fixas
        self.ws.column_dimensions['A'].width = 12  # ID
        self.ws.column_dimensions['B'].width = 45  # Título
        self.ws.column_dimensions['C'].width = 15  # Ref Feed
        self.ws.column_dimensions['D'].width = 12  # Preço Feed
        
        # Colunas por loja (começam em E)
        col_letter = 5
        for _ in self.store_names:
            self.ws.column_dimensions[get_column_letter(col_letter)].width = 12      # Preço
            self.ws.column_dimensions[get_column_letter(col_letter + 1)].width = 10  # Dif%
            self.ws.column_dimensions[get_column_letter(col_letter + 2)].width = 40  # URL
            col_letter += 3
    
    def add_product_row(self, product: FeedProduct, 
                       store_results: Dict[str, Optional[Dict]]):
        """
        Adiciona linha de produto.
        
        Args:
            product: FeedProduct do feed
            store_results: Dict {store_name: result_dict ou None}
                result_dict tem: url, price_text, price_num, confidence
        """
        # Colunas base
        row_data = [
            product.id,
            product.title,
            product.ref_raw,
            product.price_text,
        ]
        
        # Para cada loja
        for store_name in self.store_names:
            result = store_results.get(store_name)
            
            if result and result.get("price_text"):
                # Produto encontrado
                price_text = result["price_text"]
                price_num = result.get("price_num")
                url = result.get("url", "")
                
                # Calcular diferença %
                diff_value = None
                diff_display = None
                
                if product.price_num and price_num:
                    # Ambos têm valores numéricos - calcular diferença
                    diff_value = (price_num - product.price_num) / product.price_num
                    diff_display = diff_value  # Será formatado como %
                elif price_text and not price_num:
                    # NOVO: Tem texto de preço mas não conseguiu converter
                    # Provavelmente preço promocional Black Friday
                    diff_display = "N/A"  # Mostrar N/A em vez de vazio
                    diff_value = "N/A"    # Marcar para formatação amarela
                
                row_data.extend([price_text, diff_display, url])
            else:
                # Produto não encontrado
                row_data.extend(["--", None, ""])
        
        # Adicionar linha
        row_num = self.ws.max_row + 1
        self.ws.append(row_data)
        
        # Aplicar formatação
        self._format_row(row_num, len(self.store_names))
    
    def add_product(self, product: FeedProduct, 
                   store_results: Dict[str, Optional[Dict]]):
        """
        Alias para add_product_row (compatibilidade).
        """
        return self.add_product_row(product, store_results)
    
    def to_buffer(self):
        """
        Retorna Excel como BytesIO buffer (para Streamlit download).
        
        Returns:
            BytesIO buffer com Excel
        """
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def save(self, path: Path):
        """
        Salva Excel em ficheiro.
        
        Args:
            path: Caminho do ficheiro
        """
        self.wb.save(path)
    
    def _format_row(self, row_num: int, num_stores: int):
        """
        Aplica formatação a uma linha de dados.
        v4.9.1: Adiciona suporte para N/A amarelo
        
        Args:
            row_num: Número da linha
            num_stores: Número de lojas
        """
        # Colunas base (A-D)
        for col in range(1, 5):
            cell = self.ws.cell(row_num, col)
            cell.border = self.border
            cell.alignment = Alignment(vertical="center")
        
        # Título (coluna B) - wrap text
        self.ws.cell(row_num, 2).alignment = Alignment(vertical="center", wrap_text=True)
        
        # Para cada loja (3 colunas por loja)
        col = 5
        for store_idx in range(num_stores):
            # Coluna preço
            price_cell = self.ws.cell(row_num, col)
            price_cell.border = self.border
            price_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Coluna diferença %
            diff_cell = self.ws.cell(row_num, col + 1)
            diff_cell.border = self.border
            diff_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formatação da diferença
            diff_value = diff_cell.value
            
            if isinstance(diff_value, (int, float)):
                # Valor numérico - formatar como percentagem
                diff_cell.number_format = "0.0%"
                
                if diff_value > 0:
                    # Positivo = loja mais cara que tu = VERDE (ganhas)
                    diff_cell.fill = self.green_fill
                elif diff_value < 0:
                    # Negativo = loja mais barata que tu = VERMELHO (perdes)
                    diff_cell.fill = self.red_fill
            
            elif diff_value == "N/A":
                # NOVO: Preço existe mas não calculável (Black Friday)
                # Mostrar com fundo amarelo de aviso
                diff_cell.fill = self.yellow_fill
                diff_cell.font = Font(italic=True, color="C65911")  # Texto laranja escuro
            
            elif diff_value is None and price_cell.value == "--":
                # Produto não encontrado = CINZA
                price_cell.fill = self.gray_fill
                diff_cell.fill = self.gray_fill
            
            # Coluna URL
            url_cell = self.ws.cell(row_num, col + 2)
            url_cell.border = self.border
            url_cell.alignment = Alignment(vertical="center")
            
            # Se tem URL, tornar hyperlink
            if url_cell.value and url_cell.value.startswith("http"):
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="0563C1", underline="single")
                url_cell.value = "🔗 Ver produto"
            
            col += 3
    
    def freeze_header(self):
        """Congela a primeira linha (header)"""
        self.ws.freeze_panes = "A2"


# Funções auxiliares para compatibilidade

def build_excel(products: List[FeedProduct], 
               store_names: List[str], 
               all_results: Dict[str, Dict], 
               output_path: Path):
    """
    Gera Excel comparativo (compatibilidade com v4.3).
    
    Args:
        products: Lista de produtos do feed
        store_names: Lista de nomes das lojas
        all_results: Dict {store_name: {ref_norm: result_dict}}
        output_path: Caminho do Excel a gerar
    """
    builder = ExcelBuilder(store_names)
    builder._create_headers()
    
    for product in products:
        # Coletar resultados desta ref de cada loja
        store_results = {}
        for store_name in store_names:
            store_data = all_results.get(store_name, {})
            result = store_data.get(product.ref_norm)
            store_results[store_name] = result
        
        builder.add_product_row(product, store_results)
    
    builder.freeze_header()
    builder.save(output_path)


def create_single_ref_excel(ref: str, ref_norm: str, your_price: float,
                           store_names: List[str], results: List[Dict]) -> object:
    """
    Cria Excel para busca rápida de uma única referência.
    v4.9.1: Suporte para N/A amarelo
    
    Args:
        ref: Referência original
        ref_norm: Referência normalizada
        your_price: Preço do utilizador (opcional)
        store_names: Lista de lojas pesquisadas
        results: Lista de dicts com resultados
    
    Returns:
        BytesIO buffer com Excel
    """
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Busca Rápida"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Headers
    headers = ["Loja", "Preço", "Diferença", "Confiança", "URL"]
    ws.append(headers)
    
    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    
    # Larguras
    ws.column_dimensions['A'].width = 18  # Loja
    ws.column_dimensions['B'].width = 15  # Preço
    ws.column_dimensions['C'].width = 15  # Diferença
    ws.column_dimensions['D'].width = 12  # Confiança
    ws.column_dimensions['E'].width = 50  # URL
    
    # Dados
    for result in results:
        ws.append([
            result["Loja"],
            result["Preço"],
            result.get("Diferença", "N/A" if result["Preço"] not in ["Não encontrado", "—"] and result.get("Diferença") == "—" else result.get("Diferença", "")),
            result["Confiança"],
            result["URL"]
        ])
        
        row_num = ws.max_row
        
        # Aplicar borders
        for col in range(1, 6):
            cell = ws.cell(row_num, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        
        # Formatação condicional
        price_cell = ws.cell(row_num, 2)
        diff_cell = ws.cell(row_num, 3)
        
        if result["Preço"] == "Não encontrado":
            price_cell.fill = gray_fill
            diff_cell.fill = gray_fill
        elif result["Preço"].startswith("Erro"):
            price_cell.fill = red_fill
            diff_cell.fill = red_fill
        elif diff_cell.value == "N/A":
            # NOVO: Preço existe mas diferença não calculável
            diff_cell.fill = yellow_fill
            diff_cell.font = Font(italic=True, color="C65911")
        
        # URL como hyperlink
        url_cell = ws.cell(row_num, 5)
        if url_cell.value and url_cell.value.startswith("http"):
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0563C1", underline="single")
            url_cell.value = "🔗 Ver produto"
    
    # Congelar headers
    ws.freeze_panes = "A2"
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


# ============================================================================
# TESTES
# ============================================================================
if __name__ == "__main__":
    print("=== Teste de Excel Builder v4.9.1 ===\n")
    
    from .feed import FeedProduct
    from pathlib import Path
    
    # Produtos fake para teste (incluindo caso Black Friday)
    products = [
        FeedProduct(
            "001", "Escape Mivv SR-1", "https://feed.com/p1",
            "€ 331.50", 331.50,
            "H.085.LR1X", "H085LR1X", ["H085LR1X"]
        ),
        FeedProduct(
            "002", "Travão Brembo Z04", "https://feed.com/p2",
            "€ 180.00", 180.00,
            "110A26310", "110A26310", ["110A26310"]
        ),
        FeedProduct(
            "003", "Kit Transmissão DID", "https://feed.com/p3",
            "€ 150.00", 150.00,
            "KIT520VX3", "KIT520VX3", ["KIT520VX3"]
        ),
    ]
    
    # Resultados fake (incluindo preço Black Friday)
    results = {
        "wrs": {
            "H085LR1X": {"url": "https://wrs.it/p1", "price_text": "€ 365.00", "price_num": 365.0},
            "110A26310": None,  # Não encontrado
            "KIT520VX3": {"url": "https://wrs.it/p3", "price_text": "~~180€~~ 140€", "price_num": None},  # Black Friday!
        },
        "omniaracing": {
            "H085LR1X": {"url": "https://omnia.net/p1", "price_text": "€ 355.00", "price_num": 355.0},
            "110A26310": {"url": "https://omnia.net/p2", "price_text": "€ 175.00", "price_num": 175.0},
            "KIT520VX3": {"url": "https://omnia.net/p3", "price_text": "Antes: 160€ Agora: 130€", "price_num": None},
        },
    }
    
    # Gerar Excel
    output = Path("test_comparador_v491.xlsx")
    build_excel(products, ["wrs", "omniaracing"], results, output)
    
    print(f"✅ Excel gerado: {output}")
    print("\nAbre o ficheiro para ver o resultado!")
    print("Cores esperadas:")
    print("  Verde = concorrência mais cara")
    print("  Vermelho = concorrência mais barata")
    print("  Cinza = não encontrado")
    print("  AMARELO = preço Black Friday (N/A)")
